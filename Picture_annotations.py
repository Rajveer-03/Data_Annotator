'''
Important message form the Author of the Code
Kindly enter correct database details in the connection establishing module
spaces are provided for it.
Change the host if required to unless leave it as localhost for default.
'''


# importing necessary modules
import mysql.connector as c
import xlrd
import openpyxl
import tkinter as tk
from tkinter import*
from tkinter import messagebox
import urllib.request
from PIL import ImageTk, Image
from io import BytesIO
import requests
import tkinter.scrolledtext as scrolledtext

# creating connection
conn = c.connect(host = "localhost", user = """Enter User Name""", passwd = """Enter password""", database = """Enter database name""")
cur = conn.cursor()
cur.execute("create table if not exists annotator(id bigint primary key, text varchar(300), images longblob, text_annotations varchar(10), image_annotations varchar(10))")
cur = conn.cursor(buffered=True)
path = "C:\\Users\\veerm\\Desktop\\data.xlsx"

# creating function to get the data from the Excel sheet
def get_data():

    xl_loc = (path)
    xl_cur = xlrd.open_workbook(xl_loc)
    sheet = xl_cur.sheet_by_index(0)
    sheet.cell_value(0, 0)
    wrkbk = openpyxl.load_workbook(path)
    sh = wrkbk.active
    max_rows = sh.max_row
    l = []
    for i in range(1, max_rows):
        l1 = sheet.row_values(i)
        l1.append("-")
        l1.append("-")
        l1 = [i] + l1
        l.append(tuple(l1))
        
    return l

# creating a function to get the total number of rows in Excel sheet
def get_max_rows():
    xl_loc = (path)
    xl_cur = xlrd.open_workbook(xl_loc)
    sheet = xl_cur.sheet_by_index(0)
    sheet.cell_value(0, 0)
    wrkbk = openpyxl.load_workbook(path)
    sh = wrkbk.active
    return sh.max_row


# Designing the GUI usnig Tkinter module in python
class GUI():
    row_no = 1
    max_rows = get_max_rows()
    # Designing the main display
    def __init__(self, root):
        self.root = root
        self.root.title("Annotations Application")
        self.root.geometry("850x500+100+100")
        self.root.resizable(False,False)
        # Creating the master frame and the content within
        self.dataframe = Frame(self.root, bd = 10, relief = RIDGE)
        self.dataframe.place(x = 0, y = 0, width = 850, height = 500)
        self.pictureframe = Frame(self.dataframe, bd = 2, relief = RIDGE)
        self.pictureframe.place(x = 35, y = 200, width =300, height = 250)
        self.id_label = Label(self.dataframe, text = "ID: ", font =("times new roman", 12, "bold"), padx = 3, pady = 1)
        self.id_label.place(x = 75, y = 45)
        self.id_text = Text(self.dataframe, height = 1, width = 20, padx = 5, pady = 3)
        self.id_text.place(x = 105, y = 45)
        self.text_label = Label(self.dataframe, text = "Text: ", font =("times new roman", 12, "bold"), padx = 3, pady = 1)
        self.text_label.place(x = 60, y = 80)
        self.text_text = scrolledtext.ScrolledText(self.dataframe, height = 5, width = 20, padx = 5, pady = 3, wrap='word')
        self.text_text.place(x = 105, y = 80)
        

        # Creating buttons and assigning them their functionalities
        self.text_positive = Button(self.dataframe, text = "Positive", font =("times new roman", 12), height = 1, width = 10, command = self.text_positive_annotation)
        self.text_positive.place(x = 350, y= 80)
        self.text_negative = Button(self.dataframe, text = "Negative", font =("times new roman", 12), height = 1, width = 10, command = self.text_negative_annotation)
        self.text_negative.place(x = 475, y= 80)
        self.text_neutral = Button(self.dataframe, text = "Neutral", font =("times new roman", 12), height = 1, width = 10, command = self.text_neutral_annotation)
        self.text_neutral.place(x = 600, y= 80)

        self.image_positive = Button(self.dataframe, text = "Positive", font =("times new roman", 12), height = 1, width = 10, command = self.image_positive_annotation)
        self.image_positive.place(x = 350, y= 200)
        self.image_negative = Button(self.dataframe, text = "Negative", font =("times new roman", 12), height = 1, width = 10, command = self.image_negative_annotation)
        self.image_negative.place(x = 475, y= 200)
        self.image_neutral = Button(self.dataframe, text = "Neutral", font =("times new roman", 12), height = 1, width = 10, command = self.image_neutral_annotation)
        self.image_neutral.place(x = 600, y= 200)

        self.save= Button(self.dataframe, text = "Save", font =("times new roman", 12), height = 1, width = 10, command = self.save)
        self.save.place(x = 350, y= 250)
        self.nex = Button(self.dataframe, text = "Next", font =("times new roman", 12), height = 1, width = 10, command = self.next_action)
        self.nex.place(x = 350, y= 300)
        self.qui = Button(self.dataframe, text = "Quit", font =("times new roman", 12), height = 1, width = 10, command = root.destroy)
        self.qui.place(x = 350, y= 350)

        # Displaying available image in picture frame
        # else displaying a default image
        # Handeling exception using exception handling in python
        fetch = "SELECT * from annotator where id = %s"
        cur.execute(fetch,(self.row_no,))
        records = cur.fetchall()
        for row in records:
            self.id_text.insert(tk.END, row[0])
            self.text_text.insert(tk.END, row[1])
            if row[2] != b'':
                try:
                    image_url = row[2]
                    r = requests.get(image_url)
                    pilImage = Image.open(BytesIO(r.content))
                    pilImage = pilImage.resize((300, 250), Image.Resampling.LANCZOS)
                    self.image = ImageTk.PhotoImage(pilImage)
                    tk.Label(self.pictureframe, image=self.image).pack()
                except IOError as e:
                    print(e.code)
            else:
                try:
                    image_url = "https://upload.wikimedia.org/wikipedia/commons/1/14/No_Image_Available.jpg?20200913095930"
                    r = requests.get(image_url)
                    pilImage = Image.open(BytesIO(r.content))
                    pilImage = pilImage.resize((300, 250), Image.Resampling.LANCZOS)
                    self.image = ImageTk.PhotoImage(pilImage)
                    tk.Label(self.pictureframe, image=self.image).pack()
                except IOError as e:
                    print(e)
        self.text_text.config(state = DISABLED)
        self.id_text.config(state = DISABLED)

    # Assigning functionalities to the buttons
    def text_positive_annotation(self):
        sql = "UPDATE annotator SET text_annotations = 'Postive' where id = %s"
        cur.execute(sql, (self.row_no,))

    def text_negative_annotation(self):
        sql = "UPDATE annotator SET text_annotations = 'Negative' where id = %s"
        cur.execute(sql, (self.row_no,))

    def text_neutral_annotation(self):
        sql = "UPDATE annotator SET text_annotations = 'Neutral' where id = %s"
        cur.execute(sql, (self.row_no,))

    def image_positive_annotation(self):
        sql = "UPDATE annotator SET image_annotations = 'Postive' where id = %s"
        cur.execute(sql, (self.row_no,))

    def image_negative_annotation(self):
        sql = "UPDATE annotator SET image_annotations = 'Negative' where id = %s"
        cur.execute(sql, (self.row_no,))

    def image_neutral_annotation(self):
        sql = "UPDATE annotator SET image_annotations = 'Neutral' where id = %s"
        cur.execute(sql, (self.row_no,))
    
    # Assigning functionality to the save button
    def save(self):
            conn.commit()

    # creating next image action
    def next_action(self):
        if (self.row_no + 1) <= self.max_rows:
            self.text_text.config(state = NORMAL)
            self.id_text.config(state = NORMAL)
            for widget in self.pictureframe.winfo_children():
                widget.destroy()
            self.text_text.delete("1.0", "end")
            self.id_text.delete("1.0", "end")
            self.row_no += 1
            fetch = "SELECT * from annotator where id = %s"
            cur.execute(fetch,(self.row_no,))
            records = cur.fetchall()
            for row in records:
                self.id_text.insert(tk.END, row[0])
                self.text_text.insert(tk.END, row[1])
                if row[2] != b'':
                    try:
                        image_url = row[2]
                        r = requests.get(image_url)
                        pilImage = Image.open(BytesIO(r.content))
                        pilImage = pilImage.resize((300, 250), Image.Resampling.LANCZOS)
                        self.image = ImageTk.PhotoImage(pilImage)
                        tk.Label(self.pictureframe, image=self.image).pack()
                    except IOError as e:
                        print(e)
                else:
                    try:
                        image_url = "https://upload.wikimedia.org/wikipedia/commons/1/14/No_Image_Available.jpg?20200913095930"
                        r = requests.get(image_url)
                        pilImage = Image.open(BytesIO(r.content))
                        pilImage = pilImage.resize((300, 250), Image.Resampling.LANCZOS)
                        self.image = ImageTk.PhotoImage(pilImage)
                        tk.Label(self.pictureframe, image=self.image).pack()
                    except IOError as e:
                        print(e)
            self.text_text.config(state = DISABLED)
            self.id_text.config(state = DISABLED)
        else:
            messagebox.showerror('Data limit', 'Error: Data limit for the database has been reached')

# Driver Program
flag = 0
cur.execute("SELECT COUNT(*) FROM annotator")
number_of_rows = cur.fetchone()[0]
if number_of_rows == 0:
    l = get_data()
    Statement = "insert into annotator(id, text, images, text_annotations, image_annotations) values (%s, %s, %s, %s, %s)"
    cur.executemany(Statement, l)
    conn.commit()
root = tk.Tk()
obj = GUI(root)
root.mainloop()
conn.close()
